# -*- coding: UTF-8 -*-
# Copyright 2013-2018 Rumma & Ko Ltd
# License: BSD (see file COPYING for details)

"""
Database models for `lino_voga.lib.courses`.

.. xfile:: courses/Enrolment/item_description.html

     The template used to fill the items description.

"""

from __future__ import unicode_literals
from __future__ import print_function

from builtins import map
from builtins import str

from django.utils.translation import ugettext_lazy as _
from django.utils.translation import pgettext_lazy as pgettext

from lino.utils.mti import get_child
from lino.api import dd, rt

from lino.mixins import Referrable
from lino.modlib.printing.mixins import Printable
from lino_xl.lib.invoicing.mixins import InvoiceGenerator
from lino_xl.lib.courses.mixins import Enrollable
from lino_xl.lib.ledger.utils import DEBIT
from lino.utils import join_elems

from lino_xl.lib.courses.models import *

contacts = dd.resolve_app('contacts')

from lino_xl.lib.cal.utils import day_and_month

# from lino.utils.media import TmpMediaFile

from lino.modlib.printing.utils import CustomBuildMethod


"""The default activity are **courses**.  a **hike** usually includes
a bus travel. One enrolment can mean several participants (seats).  A
**journey** also includes a room in a hotel.

"""
CourseAreas.clear()
add = CourseAreas.add_item
add('C', _("Courses"), 'default')    # one place per enrolment
add('H', _("Hikes"), 'hikes', 'courses.Hikes')
add('J', _("Journeys"), 'journeys', 'courses.Journeys')


class XlsColumn(object):

    def __init__(self, label, func, width=None, **styles):
        self.label = label
        self.func = func
        self.styles = styles
        self.width = width


class XlsTable(object):
    def __init__(self):
        self.columns = []

    def add_column(self, *args, **kwargs):
        self.columns.append(XlsColumn(*args, **kwargs))

    def write_to_sheet(self, sheet, rows):
        rowno = 1
        for i, col in enumerate(self.columns):
            # sheet.write(rowno, i, label)
            cell = sheet.cell(row=rowno, column=i+1)
            cell.value = col.label
            for k, v in col.styles.items():
                setattr(cell, k, v)
            if col.width is not None:
                sheet.column_dimensions[cell.column].width = col.width
            
        for row in rows:
            rowno += 1
            for i, col in enumerate(self.columns):
                value = col.func(row)
                # sheet.write(rowno, i, value)
                sheet.cell(row=rowno, column=i+1).value = value


class CourseToXls(CustomBuildMethod):
    """Interesting, but currently not used."""
    target_ext = '.xlsx'
    name = 'course2xls'
    label = _("Export")

    def custom_build(self, ar, obj, target):
        from openpyxl.workbook import Workbook
        from openpyxl.styles import Alignment
        events = obj.events_by_course().order_by('start_date')

        xt = XlsTable()
        
        # def func(enr):
        #     # s = ''.join([str(e) for e in enr.pupil_info])
        #     s = enr.pupil_info.text
        #     print(20160512, s, E.tostring(enr.pupil_info))
        #     return s
        # xt.add_column("Teilnehmer", func)
        xt.add_column(
            "Teilnehmer", lambda enr: enr.pupil_info.text,
            alignment=Alignment(
                horizontal="general", vertical="top",
                wrap_text=True))

        xt.add_column("Anzahl", lambda enr: enr.places)
        xt.add_column("Start", lambda enr: enr.start_date)
        xt.add_column("End", lambda enr: enr.end_date)
        xt.add_column("Invoicing", lambda enr: enr.invoicing_info.text)
        xt.add_column("Payment", lambda enr: enr.payment_info)

        for i, evt in enumerate(events):
            lbl = day_and_month(evt.start_date)

            def func(enr):
                qs = rt.models.cal.Guest.objects.filter(
                    event=evt, partner=enr.pupil)
                n = qs.count()
                if n == 0:
                    return ''
                return n
                
            xt.add_column(
                lbl, func, alignment=Alignment(
                    vertical="center", text_rotation=90), width=4)
        xt.add_column("...", lambda enr: "")
        xt.add_column("...", lambda enr: "")
        xt.add_column("...", lambda enr: "")

        wb = Workbook()
        # sheet = wb.add_sheet(str(obj))
        ws = wb.active
        ws.title = str(obj)
        # ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        # print(type(six.text_type('100')))
        ws.column_dimensions["A"].width = 40
        ws.row_dimensions[1].height = 30
        xt.write_to_sheet(ws, obj.enrolments)
        wb.save(target)


class TeacherType(Referrable, mixins.BabelNamed, Printable):

    class Meta:
        app_label = 'courses'
        abstract = dd.is_abstract_model(__name__, 'TeacherType')
        # verbose_name = _("Teacher type")
        # verbose_name_plural = _('Teacher types')
        verbose_name = _("Instructor Type")
        verbose_name_plural = _("Instructor Types")



class Teacher(contacts.Person):
    """A **teacher** is a person with an additional field
    :attr:`teacher_type`.

    .. attribute:: teacher_type

        Pointer to :class:`TeacherType`.

    """
    class Meta:
        app_label = 'courses'
        abstract = dd.is_abstract_model(__name__, 'Teacher')
        verbose_name = _("Instructor")
        verbose_name_plural = _("Instructors")

        # verbose_name = _("Teacher")
        # verbose_name_plural = _("Teachers")

    teacher_type = dd.ForeignKey('courses.TeacherType', blank=True, null=True)

    def __str__(self):
        return self.get_full_name(salutation=False)


class PupilType(Referrable, mixins.BabelNamed, Printable):

    class Meta:
        app_label = 'courses'
        abstract = dd.is_abstract_model(__name__, 'PupilType')
        # verbose_name = _("Pupil type")
        # verbose_name_plural = _('Pupil types')
        verbose_name = _("Participant Type")
        verbose_name_plural = _("Participant Types")


class Pupil(Enrollable, contacts.Person):
    """A **pupil** is a person with an additional field
    :attr:`pupil_type`.

    .. attribute:: pupil_type

        Pointer to :class:`PupilType`.

    """

    class Meta:
        app_label = 'courses'
        abstract = dd.is_abstract_model(__name__, 'Pupil')
        verbose_name = _("Participant")
        verbose_name_plural = _("Participants")
        # verbose_name = _("Pupil")
        # verbose_name_plural = _("Pupils")

    pupil_type = dd.ForeignKey('courses.PupilType', blank=True, null=True)

    # suggested_courses = dd.ShowSlaveTable('courses.SuggestedCoursesByPupil')

    def get_enrolment_info(self):
        if self.pupil_type:
            return self.pupil_type.ref

    @classmethod
    def setup_parameters(cls, fields):
        fields.update(
            partner_list=dd.ForeignKey(
                'lists.List', blank=True, null=True))

        super(Pupil, cls).setup_parameters(fields)

    @classmethod
    def get_request_queryset(cls, ar):
        qs = super(Pupil, cls).get_request_queryset(ar)
        pv = ar.param_values
        if pv.course:
            qs = qs.filter(
                Q(enrolments_by_pupil__start_date__isnull=True) |
                Q(enrolments_by_pupil__start_date__lte=dd.today()))
            qs = qs.filter(
                Q(enrolments_by_pupil__end_date__isnull=True) |
                Q(enrolments_by_pupil__end_date__gte=dd.today()))
            qs = qs.distinct()
            # qs = qs.filter(enrolments_by_pupil__course=pv.course)
            # qs = qs.filter(
            #     enrolments_by_pupil__state__in=EnrolmentStates.filter(
            #         invoiceable=True))
            qs = qs.filter(
                enrolments_by_pupil__course=pv.course,
                enrolments_by_pupil__state__in=EnrolmentStates.filter(
                    invoiceable=True))
            # qs = qs.filter(
            #     enrolments_by_pupil__state=EnrolmentStates.confirmed)
            

        if pv.partner_list:
            qs = qs.filter(list_memberships__list=pv.partner_list)
        return qs

    @classmethod
    def get_title_tags(self, ar):
        for t in super(Pupil, self).get_title_tags(ar):
            yield t
        pv = ar.param_values
        if pv.partner_list:
            yield str(pv.partner_list)


# class CreateInvoicesForCourse(CreateInvoice):
#     """
#     Create invoices for all participants of this course.
#     """
#     def get_partners(self, ar):
#         course = ar.selected_rows[0]
#         return [obj.pupil for obj in course.enrolment_set.filter(
#             state=EnrolmentStates.confirmed)]


class CourseType(Referrable, mixins.BabelNamed):

    class Meta:
        app_label = 'courses'
        abstract = dd.is_abstract_model(__name__, 'CourseType')
        verbose_name = _("Activity type")
        verbose_name_plural = _('Activity types')


class Line(Line):

    class Meta(Line.Meta):
        app_label = 'courses'
        abstract = dd.is_abstract_model(__name__, 'Line')

    course_type = dd.ForeignKey('courses.CourseType', blank=True, null=True)



class Course(Referrable, Course):
    """Extends the standard model by adding a field :attr:`fee`.

    Also adds a :attr:`ref` field and defines a custom :meth:`__str__`
    method.

    The custom :meth:`__str__` method defines how to textually
    represent a course e.g. in the dropdown list of a combobox or in
    reports. Rules:

    - If :attr:`ref` is given, it is shown, but see also the two
      following cases.

    - If :attr:`name` is given, it is shown (possibly behind the
      :attr:`ref`).

    - If a :attr:`line` (series) is given, it is shown (possibly
      behind the :attr:`ref`).

    - If neither :attr:`ref` nor :attr:`name` nor :attr:`line` are
      given, show a simple "Course #".


    .. attribute:: ref
    
        An identifying public course number to be used by both
        external and internal partners for easily referring to a given
        course.

    .. attribute:: name

        A short designation for this course. An extension of the
        :attr:`ref`.

    .. attribute:: line

        Pointer to the course series.


    .. attribute:: fee

        The default participation fee to apply for new enrolments.

    .. attribute:: payment_term

        The payment term to use when writing an invoice. If this is
        empty, Lino will use the partner's default payment term.

    .. attribute:: paper_type

        The paper_type to use when writing an invoice. If this is
        empty, Lino will use the site's default paper type.

    """
    class Meta(Course.Meta):
        # app_label = 'courses'
        abstract = dd.is_abstract_model(__name__, 'Course')
        verbose_name = _("Activity")
        verbose_name_plural = _('Activities')

    fee = dd.ForeignKey('products.Product',
                        blank=True, null=True,
                        verbose_name=_("Default participation fee"),
                        related_name='courses_by_fee')

    payment_term = dd.ForeignKey(
        'ledger.PaymentTerm',
        related_name="%(app_label)s_%(class)s_set_by_payment_term",
        blank=True, null=True)    

    paper_type = dd.ForeignKey(
        'sales.PaperType',
        related_name="%(app_label)s_%(class)s_set_by_paper_type",
        blank=True, null=True)    

    quick_search_fields = 'name line__name line__topic__name ref'

    # course2xls = CourseToXls.create_action()

    @classmethod
    def get_registrable_fields(cls, site):
        for f in super(Course, cls).get_registrable_fields(site):
            yield f
        yield 'fee'

    @dd.chooser()
    def fee_choices(cls, line):
        Product = rt.models.products.Product
        if not line or not line.fees_cat:
            return Product.objects.none()
        return Product.objects.filter(cat=line.fees_cat)

    def __str__(self):
        if self.name:
            if self.ref:
                return "{0} {1}".format(self.ref, self.name)
            return self.name
        if self.ref:
            if self.line:
                return "{0} {1}".format(self.ref, self.line)
            return self.ref
        # Note that we cannot use super() with
        # python_2_unicode_compatible
        return "{0} #{1}".format(self._meta.verbose_name, self.pk)

    def update_cal_summary(self, et, i):
        label = dd.babelattr(et, 'event_label')
        if self.ref:
            label = self.ref + ' ' + label
        return "%s %d" % (label, i)

Course.set_widget_options('ref', preferred_with=6)

# class CreateInvoiceForEnrolment(CreateInvoice):

#     def get_partners(self, ar):
#         return [o.pupil for o in ar.selected_rows]


class Enrolment(Enrolment, InvoiceGenerator):
    """Adds

    .. attribute:: fee

        The participation fee to apply for this enrolment.

    .. attribute:: free_events

        Number of events to add for first invoicing for this
        enrolment.

    .. attribute:: amount

        The total amount to pay for this enrolment. This is
        :attr:`places` * :attr:`fee`.

    .. attribute:: pupil_info

        Show the name and address of the participant.  Overrides
        :attr:`lino_xl.lib.courses.models.Enrolment.pupil_info`
        in order to add (between parentheses after the name) some
        information needed to compute the price.

    .. attribute:: invoicing_info

        A virtual field showing a summary of recent invoicings.

    .. attribute:: payment_info

        A virtual field showing a summary of due accounting movements
        (debts and payments).

    """

    # invoiceable_date_field = 'request_date'
    _invoicing_info = None

    class Meta(Enrolment.Meta):
        abstract = dd.is_abstract_model(__name__, 'Enrolment')
        # abstract = False  # dd.is_abstract_model(__name__, 'Enrolment')
        verbose_name = _("Enrolment")
        verbose_name_plural = _("Enrolments")
        # in Voga it is allowed to enrole several times
        unique_together = None

    amount = dd.PriceField(_("Amount"), blank=True, null=True)

    fee = dd.ForeignKey('products.Product',
                        blank=True, null=True,
                        # verbose_name=_("Participation fee"),
                        related_name='enrolments_by_fee')

    free_events = models.IntegerField(
        pgettext("in an enrolment", "Free events"),
        null=True, blank=True,
        help_text=_("Number of events to add for first invoicing "
                    "for this enrolment."))

    # create_invoice = CreateInvoiceForEnrolment()

    def get_invoiceable_partner(self):
        # if hasattr(self.pupil, 'salesrule'):
        #     return self.pupil.salesrule.invoice_recipient or self.pupil
        return self.pupil

    def get_invoiceable_payment_term(self):
        return self.course.payment_term

    def get_invoiceable_paper_type(self):
        return self.course.paper_type
        # if hasattr(self.pupil, 'salesrule'):
        #     return self.pupil.salesrule.paper_type

    def get_invoiceable_events(self, start_date, max_date):
        flt = dict(
            start_date__lte=max_date,
            state=rt.models.cal.EntryStates.took_place)
        if start_date:
            flt.update(start_date__gte=start_date)
        return self.course.events_by_course(**flt).order_by('start_date')
            
    def get_invoiceable_free_events(self):
        return self.free_events

    @classmethod
    def get_generators_for_plan(cls, plan, partner=None):
        """
        Yield all enrolments for which the given plan and partner should
        generate an invoice.
        """
        qs = cls.objects.all()
        # **{
        #     cls.invoiceable_date_field + '__lte': plan.max_date or plan.today})
        if plan.course is None:
            qs = qs.filter(course__state=CourseStates.active)
        else:
            qs = qs.filter(course__id=plan.course.id)
        if partner is None:
            partner = plan.partner
        if partner:
            pupil = get_child(partner, rt.models.courses.Pupil)
            # pupil = partner.get_mti_child('pupil')
            if pupil:  # isinstance(partner, rt.models.courses.Pupil):
                q1 = models.Q(
                    pupil__salesrule__invoice_recipient__isnull=True, pupil=pupil)
                q2 = models.Q(pupil__salesrule__invoice_recipient=partner)
                qs = qs.filter(models.Q(q1 | q2))
            else:
                # if the partner is not a pupil, then it might still
                # be an invoice_recipient
                qs = qs.filter(pupil__salesrule__invoice_recipient=partner)
                
        # dd.logger.info("20160513 %s (%d rows)", qs.query, qs.count())
        return qs.order_by('id')

    @dd.chooser()
    def fee_choices(cls, course):
        Product = rt.models.products.Product
        if not course or not course.line or not course.line.fees_cat:
            return Product.objects.none()
        return Product.objects.filter(cat=course.line.fees_cat)

    def full_clean(self, *args, **kwargs):
        if self.fee_id is None and self.course_id is not None:
            self.fee = self.course.fee
            if self.fee_id is None and self.course.line_id is not None:
                self.fee = self.course.line.fee
        # if self.number_of_events is None:
        #     if self.fee_id and self.fee.number_of_events:
        #         self.number_of_events = self.fee.number_of_events
        #     self.number_of_events = self.course.max_events
        if self.amount is None:
            self.compute_amount()
        super(Enrolment, self).full_clean(*args, **kwargs)

    def pupil_changed(self, ar):
        self.compute_amount()

    def places_changed(self, ar):
        self.compute_amount()

    # def fee_changed(self, ar):
    #     if self.fee_id is not None:
    #         self.number_of_events = self.fee.number_of_events
    #     self.compute_amount()

    # def get_number_of_events(self):
    #     if self.number_of_events is not None:
    #         return self.number_of_events
    #     if self.fee_id and self.fee.number_of_events:
    #         return self.fee.number_of_events
    #     return self.course.max_events or 0

    def get_invoiceable_amount(self):
        return self.amount

    def get_invoiceable_product(self, max_date=None):
        return self.fee
    
    def compute_amount(self):
        #~ if self.course is None:
            #~ return
        if self.places is None:
            return
        if self.fee is None:
            return
        # When `products` is not installed, then fee may be None
        # because it is a DummyField.
        price = getattr(self.fee, 'sales_price') or ZERO
        try:
            self.amount = price * self.places
        except TypeError as e:
            logger.warning("%s * %s -> %s", price, self.places, e)

    def get_invoiceable_title(self, number=None):
        title = _("{enrolment} to {course}").format(
            enrolment=self.__class__._meta.verbose_name,
            course=self.course)
        # if self.fee.tariff and self.fee.tariff.number_of_events:
        #     info = self.compute_invoicing_info()
        #     number = info.invoice_number(invoice)
        if number is None:
            return title
        if number > 1:
            msg = _("[{number}] Renewal {title}")
        else:
            msg = _("[{number}] {title}")
        return msg.format(title=title, number=number)

    def get_invoiceable_start_date(self, max_date):
        return self.start_date or self.course.start_date
    
    def get_invoiceable_qty(self):
        return self.places

    def setup_invoice_item(self, item):
        item.description = dd.plugins.jinja.render_from_request(
            None, 'courses/Enrolment/item_description.html',
            obj=self, item=item)

    def get_invoiceable_info(self, plan):
        """Return the product to use for the invoice.
        This also decides whether an invoice should be issued or not.
        """
        # dd.logger.info('20160223 %s', self.course)
        if not self.course.state.is_invoiceable:
            return
        if not self.state.invoiceable:
            return
        max_date = plan.max_date or plan.today

        # the following 2 lines were nonsense. it is perfectly okay to
        # write an invoice for an enrolment which starts in the
        # future.
        # if self.start_date and self.start_date > max_date:
        #     return

        # but at least for our demo fixtures we don't want invoices
        # for enrolments in the future:
        if self.request_date and self.request_date > max_date:
            return

        return self.compute_invoicing_info(max_date)

    @dd.virtualfield(dd.HtmlBox(_("Participant")))
    def pupil_info(self, ar):
        if not self.pupil_id:
            return ''
        elems = []
        txt = self.pupil.get_full_name(nominative=True)
        if ar is None:
            elems = [txt]
        else:
            elems = [ar.obj2html(self.pupil, txt)]
        info = self.pupil.get_enrolment_info()
        if info:
            # elems += [" ({})".format(self.pupil.pupil_type.ref)]
            elems += [" ({})".format(info)]
        elems += [', ']
        elems += join_elems(
            self.pupil.address_location_lines(), sep=', ')
        if self.pupil.phone:
            elems += [', ', _("Phone: {0}").format(self.pupil.phone)]
        if self.pupil.gsm:
            elems += [', ', _("GSM: {0}").format(self.pupil.gsm)]
        return E.p(*elems)

    @dd.displayfield(_("Payment info"))
    def payment_info(self, ar):
        if not self.pupil_id:
            return ''
        return rt.models.ledger.Movement.balance_info(
            DEBIT, partner=self.pupil, cleared=False)
        

# dd.inject_field(
#     'products.Product', 'number_of_events',
#     models.IntegerField(
#         _("Number of events"), null=True, blank=True,
#         help_text=_("Number of events paid per invoicing.")))

# dd.inject_field(
#     'products.Product', 'min_asset',
#     models.IntegerField(
#         _("Invoice threshold"), null=True, blank=True,
#         help_text=_("Minimum number of events to pay in advance.")))

